import os
import glob
import shutil
import time
# import portalocker
import logging
from getpass import getuser
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.colors import RED, GREEN
"""
Category and category group classes
Keynote and keynote group classes
File I/O

Spreadsheet file is the master file and is used for loading.
.txt file can be saved, as can the spreadsheet.
"""
logger = logging.getLogger('__name__')
stream_handler = logging.StreamHandler()
logger.addHandler(stream_handler)


class Category(object):
    """
    Information for a specific category.
    A Category holds three dictionaries of keynotes.
    """

    def __init__(self, number, name):
        """
        Create category object
        """
        logger.debug("Creating category {}".format(name))
        self.name = name
        self.number = int(number)
        # Placeholders for keynotes to be added later
        self.demoKeynotes = []
        self.existingKeynotes = []
        self.newKeynotes = []

    def addKeynote(self, keynote):
        """
        Add a keynote into the correct category
        """
        if keynote.den == 'D':
            self.demoKeynotes.append(keynote)
        elif keynote.den == 'E':
            self.existingKeynotes.append(keynote)
        else:
            self.newKeynotes.append(keynote)
        if not keynote.number:
            logger.error('Keynote number not provided.')
        keynote.catnum = self.number
        keynote.category = self

    @property
    def keynotes(self):
        return self.demoKeynotes + self.existingKeynotes + self.newKeynotes

    @property
    def fullstring(self):
        return f"{self.number}\t{self.name}"

    def __repr__(self):
        return f"Category({self.name}, {self.number})"

    def __str__(self):
        return f"Category {self.number}: {self.name}"


class Keynote(object):
    """
    Information for a single keynote.
    """

    def __init__(self, number=-1, kType=None, disabled=False, category=None,
                 numString=None, kText='<Empty>', catString=None):
        """
        Build a Keynote.
        """

        def _from_file_data(self, num_string, text_string, cat_string):
            """
            Build based on input from a file.
            """
            if len(num_string) == 5:
                if num_string[0] in 'DEN':
                    self.den = num_string[0]   # First character is D, E, or N
                    try:
                        self.catnum = int(num_string[1:3])  # Category Number
                        self.number = int(num_string[3:6])
                    except Exception:
                        logging.error(f"Error in keynote {num_string}")
                        return False
                    self.category = None
                else:
                    logging.error(f"Bad first character: <{numString}>")
            else:
                logging.error(f"Bad number string: <{numString}>")
            self.text = text_string
            self.disabled = cat_string == 'disabled'

        def _insert_new(self, category, den):
            """
            Build a new keynote for a known category and type (DEN)
            """
            def _next_num(keynote_group):
                if keynote_group == []:
                    return 1
                else:
                    return keynote_group[-1].number + 1

            self.text = '<Empty>'
            self.disabled = False
            self.category = category
            self.catnum = category.number
            self.den = den
            if den == 'D':
                self.number = _next_num(category.demoKeynotes)
            elif den == 'E':
                self.number = _next_num(category.existingKeynotes)
            elif den == 'N':
                self.number = _next_num(category.newKeynotes)
            else:
                logger.error(f'Bad keynote type: {den}')

        if numString and kText and catString:
            _from_file_data(self, numString, kText, catString)
        elif kType and category:
            _insert_new(self, category, kType)
        else:
            logger.error('Keynote improper arguments')
            return
        self.filter = False  # Keynotes are not initially filtered out
        # logger.debug(f'Created keynote: {self.fullstring}')

    @property
    def identifier(self):
        return "{}{:02d}{:02d}".format(self.den, self.catnum, self.number)

    @property
    def fullstring(self):
        if self.disabled:
            d = 'disabled'
        elif self.category:
            d = self.category.number
        else:
            d = '-'
        return f"{self.identifier}\t{self.text}\t{d}"

    def __str__(self):
        self.fullstring

    def __repr__(self):
        if self.disabled:
            cat = 'disabled'
        else:
            cat = self.catnum
        return (f"Keynote(numString={self.identifier}, "
                "kText='{self.text}', catString={cat})")


class keynoteFile(object):
    """
    A record containing all the data for a keynote file.
    """

    def __init__(self):
        self.fileName = None   # Store the original filename
        self.user = getuser()  # We use the username for the lock
        self.modified = False  # Tells us if the file content has been changed
        self.categories = []   # A list of categories, with keynotes attached
        self.lockError = ""    # A string explaining why a file can't be locked
        logger.debug('Created new empty keynoteFile record')

    def clear(self):
        """
        Delete the contents of a loaded file.
        """
        self.categories = []
        self.fileName = None
        self.modified = False

    def lockedName(self, name):
        parts = name.rsplit('.', 1)
        return parts[0] + '_' + self.user + '.' + parts[1]

    def lockable(self, name):
        """
        Determine if the file is present and can be locked. No side effects.
        """
        logger.debug(f'Checking lock on file: {name}')
        if not os.path.isfile(name):
            self.lockError = "File does not exist."
            return False
        # Check for an Excel lock file
        dir, file = os.path.split(name)
        lock_file = f"~${file}"
        lock_path = os.path.join(dir, lock_file)
        if os.path.isfile(lock_path):
            self.lockError = "File is locked by Excel"
            return False
        # Check it's not locked by knm
        try:
            user = file.upper().split("_")[1].split(".XLSX")[0]
        except IndexError:
            logger.debug("File is not locked by a user")
        else:
            self.lockError = f'File is locked by user {user}'
            return False
        # Check that the file is named properly
        if name.upper().count('NOTES.XLSX') == 0:
            self.lockError = "Not a keynote file"
            return False
        return True

    def lockFile(self, name):
        """
        Make a backup then lock the file (via copy)
        """
        if self.lockable(name):
            logger.debug(f'Locking file: {name}')
            # Execute the backup and lock
            try:
                # Make a backup, after deleting existing backups
                backups = glob.glob(name + '.[0-9][0-9][0-9]*')
                logger.debug(f'Backups found: {backups}')
                if backups:
                    for f in backups:
                        logger.debug(f'Removing {f}')
                        os.remove(f)
                # Make the new backup
                shutil.copy(name, name + '.' + str(int(time.time())))
                os.rename(name, self.lockedName(name))
                self.fileName = name
            except Exception as e:
                logger.warn("Unable to lock file")
                logger.error(str(e))
                return False
        return name  # Will be False if there's no file

    def unlockFile(self, name):
        logger.debug(f'Unlocking file {name}')
        result = os.path.isfile(self.lockedName(name))
        if result:
            os.rename(self.lockedName(name), name)
        else:
            logger.error(f'File not found: {name}')
        self.clear()
        return result  # Can't unlock non-existent file

    def checkLock(self, fileName=None):
        """
        Before opening or saving a file, create and/or release a lock.
        Save check is implemented through the GUI.
        """
        if self.fileName is not None:  # We have a file already open
            # Unlock self.fileName
            return self.unlockFile(fileName)
        if fileName is not None:
            # Lock the new file
            return self.lockFile(fileName)

    def load(self, fileName, fileType):
        """
        Load the new file. Locks only apply to the spreadsheet.
        """
        # if fileType == 'Text':
        #     self.loadTxt(self.lockedName(fileName))
        if fileType == 'Excel':
            if self.checkLock(fileName):
                return self.loadXlsx(self.lockedName(fileName))
            else:
                logger.warn(f'File not found: {fileName}')
        else:
            logger.warn('Bad fileType: {fileType}')
        return False

    # def loadTxt(self, keynoteFile):
    #     """
    #     Load in a file full of keynotes and return categories and keynotes.
    #     """
    #     # Clear out any existing categories or keynotes
    #     self.categories = []
    #     keynoteList = []
    #     # Load in the new stuff
    #     with open(keynoteFile, "r") as f:
    #         # Save the filename now that we know it opened OK
    #         self.fileName = keynoteFile
    #         line = f.readline()
    #         while line != '':  # Empty string means end of file
    #             line = line.rstrip('\t \n')
    #             ll = line.split('\t')
    #             # print(ll)
    #             if len(ll) == 2:
    #                 self.categories.append(Category(ll[0], ll[1]))
    #             elif len(ll) == 3:
    #                 keynoteList.append(Keynote(numString=ll[0], kText=ll[1],
    #                                            catString=ll[2]))
    #             else:
    #                 print("Ignoring line <{}>".format(line))
    #             line = f.readline()
    #
    #         print("{} categories found.".format(len(self.categories)))
    #         print("{} keynotes found".format(len(keynoteList)))
    #         for c in self.categories:
    #             # Attach keynotes to categories correctly
    #             for k in keynoteList:
    #                 if k.catnum == c.number:
    #                     k.category = c
    #                     c.addKeynote(k)
    #     # self.pprint()
    #     return self.categories

    def saveTxt(self):
        """
        Write out the keynote data in the record to a tab - delimited file.
        """
        # Make sure the filename ends in .xlsx, changing it if necessary
        txtFileName = self.fileName.rsplit('.', 1)[0] + '.txt'
        logger.info(f'Saving file: {txtFileName}')
        with open(txtFileName, 'w+') as f:
            cCount = 0
            kCount = 0
            for c in self.categories:
                f.write(f"{c.number}\t{c.name}")
                f.write('\n')
                cCount += 1
            f.write("disabled\tdisabled\n")
            f.write('\n')  # A blank line
            for c in self.categories:
                for k in c.keynotes:
                    f.write(k.fullstring)
                    f.write('\n')
                    kCount += 1
                f.write('\n')  # A blank line
            return (cCount, kCount)

    def loadXlsx(self, fileName):
        """
        Load in an Excel keynote fileand return categories and keynotes.
        """
        try:
            logger.debug(f"Opening file: {fileName}")
            self.workbook = load_workbook(filename=fileName)
        except Exception as e:
            logger.error(f'Excel open failed: {e}')
            return False
        # Assume first tab holds the Data
        rows = tuple(self.workbook.active.rows)
        keynoteList = []
        for row in rows:
            if not ((row[0].value or row[0].value == 0) and row[1].value):
                continue  # Ignore any row with a blank in columns A or B
            if type(row[0].value) == int:
                # It's a category
                if row[1].value == 'disabled':
                    # Ignore the disabled pseudo category
                    logger.debug("Ignoring disabled pseudo-category")
                    pass
                logger.debug(f'Found category {row[1].value}')
                self.categories.append(Category(row[0].value, row[1].value))
            elif row[0].value and len(row[0].value) == 5:
                # There's a keynote identifier
                keynoteList.append(Keynote(numString=row[0].value,
                                           kText=row[1].value,
                                           catString=str(row[2].value)))
            else:
                logger.error("Can't process /{}/{}/{}/".format(row[0].value,
                                                               row[1].value,
                                                               row[2].value))
        logger.info(f"{len(self.categories)} categories found.")
        logger.info(f"{len(keynoteList)} keynotes found")
        for c in self.categories:
            # Attach keynotes to categories correctly
            for k in keynoteList:
                if k.catnum == c.number:
                    k.category = c
                    c.addKeynote(k)
        # self.pprint()
        return self.categories

    def saveXlsx(self):
        """
        Write out the keynote data in the record to a spreadsheet.
        Assumes any updates to the record have already been made.
        """
        def writeCell(worksheet, row, col, val,
                      font=None, fill=None, align=None):
            """
            row is a number, col is a letter
            """
            addr = f'{col}{row}'
            logger.debug(f'Writing {val} into cell {addr}')
            worksheet[addr] = val
            if font:
                worksheet[addr].font = font
            if fill:
                worksheet[addr].fill = fill
            if align:
                worksheet[addr].alignment = align

        # Make sure the filename ends in .xlsx, changing it if necessary
        fileName = self.lockedName(self.fileName)
        logger.debug(f'Saving to {fileName}')
        # Create a new workbook and worksheet
        wb = Workbook()
        ws = wb.active
        # Set up widths, colors and fonts
        ws.column_dimensions['B'].width = 100
        wrap = Alignment(wrapText=True)
        boldFont = Font(b=True)
        redFont = Font(color=RED)
        greenFont = Font(color=GREEN)
        grayFont = Font(color='888888', size=9)
        grayFill = PatternFill("solid", fgColor='BBBBBB')
        # Counters to keep track of things
        rowCount = 1
        cCount = 0
        kCount = 0
        # Loop through writing categories
        for c in self.categories:
            # Write a category record
            writeCell(ws, rowCount, 'A', c.number)
            writeCell(ws, rowCount, 'B', c.name)
            cCount += 1
            rowCount += 1
        writeCell(ws, rowCount, 'A',
                  ("REMEMBER TO SAVE after editing, then SAVE FILE AS "
                   "Text (Tab delimited)(*.txt), then load/reload your "
                   "keynotes on your project Revit file so Revit can see "
                   "the changes. All keynote / text editing shall be "
                   "on the Excel file only."), font=redFont, align=wrap)
        ws.merge_cells(start_row=rowCount, end_row=rowCount,
                       start_column=1, end_column=2)

        rowCount += 2
        # Now write keynotes
        for c in self.categories:
            writeCell(ws, rowCount, 'A', c.name.upper(), font=boldFont)
            # startRow = rowCount  # First row in outline
            rowCount += 1
            for kt in (c.demoKeynotes, c.existingKeynotes, c.newKeynotes):
                for k in kt:
                    id = k.identifier
                    if id[0] == 'D':
                        f = redFont
                    elif id[0] == 'E':
                        f = None
                    else:
                        f = greenFont
                    writeCell(ws, rowCount, 'A', id, font=f)
                    writeCell(ws, rowCount, 'B', k.text, align=wrap)
                    if k.disabled:
                        writeCell(ws, rowCount, 'C', 'disabled')
                    else:
                        writeCell(ws, rowCount, 'C', k.catnum)
                    kCount += 1
                    rowCount += 1
                # endRow = rowCount
                rowCount += 1  # Leave a blank line
                writeCell(ws, rowCount, 'A', "DO NOT USE THIS ROW, "
                          "INSERT NEW ROW AS NEEDED",
                          font=grayFont, fill=grayFill)
                ws.merge_cells(start_row=rowCount, end_row=rowCount,
                               start_column=1, end_column=3)
                # Create outline for current category
                rowCount += 1
        logger.debug(f'Saving file: {fileName}')
        wb.save(fileName)
        return (cCount, kCount)

    def pprint(self):
        """
        Print a human readable version of the keynoteFile record.
        """
        for c in self.categories:
            print(c.fullstring)
            for k in c.keynotes:
                print(f"- {k.fullstring}")
